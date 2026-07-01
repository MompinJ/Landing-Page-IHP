#!/usr/bin/env python3
# Genera el Excel de costos del HUB Digital IHP (infraestructura Azure).
# Editable: los totales son formulas; al cambiar un precio se recalcula todo.

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# --- Paleta Hutchison Ports ---
SEA = "002E6D"
SKY = "009BDE"
HORIZON = "9ACAEB"
MIST = "F3F7FB"
SKY08 = "EAF6FC"
AQUA = "54BBAB"
SUNRAY = "FFC627"
WHITE = "FFFFFF"
INK = "002E6D"

f_title = Font(name="Calibri", size=18, bold=True, color=SEA)
f_sub = Font(name="Calibri", size=10, color="44597D", italic=True)
f_h = Font(name="Calibri", size=10, bold=True, color=WHITE)
f_b = Font(name="Calibri", size=10, color=INK)
f_bold = Font(name="Calibri", size=10, bold=True, color=SEA)
f_money = Font(name="Calibri", size=10, color=INK)
f_total = Font(name="Calibri", size=12, bold=True, color=WHITE)
f_kpi = Font(name="Calibri", size=22, bold=True, color=WHITE)
f_kpi_k = Font(name="Calibri", size=10, bold=True, color=WHITE)

fill_head = PatternFill("solid", fgColor=SEA)
fill_mist = PatternFill("solid", fgColor=MIST)
fill_sky08 = PatternFill("solid", fgColor=SKY08)
fill_total = PatternFill("solid", fgColor=SKY)
fill_kpi_min = PatternFill("solid", fgColor=SKY)
fill_kpi_norm = PatternFill("solid", fgColor=SEA)
fill_sunray = PatternFill("solid", fgColor=SUNRAY)

thin = Side(style="thin", color="D8E3EF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
right = Alignment(horizontal="right", vertical="center")

MONEY = '"$"#,##0;[Red]-"$"#,##0'

wb = Workbook()

# ============================================================
# HOJA 2: SERVICIOS (la construimos primero para referenciarla)
# ============================================================
ws = wb.active
ws.title = "Servicios"
ws.sheet_view.showGridLines = False

ws["A1"] = "HUB Digital IHP - Que rentamos en la nube (Azure)"
ws["A1"].font = f_title
ws["A2"] = "Precios base aproximados de lista, USD/mes, region Mexico Central. Editables: cambia la columna 'Precio unit.' y los totales se recalculan."
ws["A2"].font = f_sub
ws.merge_cells("A1:L1")
ws.merge_cells("A2:L2")

headers = [
    ("#", 4),
    ("Servicio (en simple)", 22),
    ("Producto Azure", 22),
    ("Para que sirve", 30),
    ("Que rentamos (config minima)", 26),
    ("Precio unit.\n(USD/mes)", 11),
    ("Cant.\n1 local", 7),
    ("Subtotal\nahorro", 11),
    ("Cant.\n2 locales", 8),
    ("Subtotal\nnormal", 11),
    ("Como se paga", 14),
    ("A considerar", 40),
]
HEADER_ROW = 4
for c, (txt, w) in enumerate(headers, start=1):
    cell = ws.cell(row=HEADER_ROW, column=c, value=txt)
    cell.font = f_h
    cell.fill = fill_head
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
    ws.column_dimensions[get_column_letter(c)].width = w

# Datos: (servicio, producto, para_que, que_rentamos, precio, cant1, cant2, pago, considerar)
rows = [
    ("El local de las apps", "Azure App Service (Plan P1v3 Linux)",
     "Mantiene encendidos los programas (recepcion y 3 areas) las 24 horas.",
     "1 plan al inicio; 2 recomendado (recepcion aparte)", 120, 1, 2, "Fijo mensual",
     "Pieza mas cara. Es el minimo que soporta red privada (no baja de P1v3). Se puede reservar 1-3 anos (-30% a -55%). El 2o local es el salto de costo mas grande."),
    ("Memoria rapida", "Azure Cache for Redis (Standard C1, 1 GB)",
     "Acelera la plataforma y guarda las sesiones de usuario.",
     "Standard C1 (con replica de respaldo)", 100, 1, 1, "Fijo mensual",
     "2a mas cara. Standard incluye replica y garantia (SLA). Existe Basic C1 (~$40) sin respaldo si se busca ahorrar."),
    ("La entrada", "Azure Front Door (Standard)",
     "Puerta unica al publico: filtra, protege y acelera las visitas.",
     "Plan Standard", 35, 1, 1, "Base + uso",
     "Cobra base mensual + visitas + datos transferidos. Premium (con firewall WAF) cuesta desde ~$300/mes."),
    ("El archivero", "Azure Database for PostgreSQL (Flexible B1ms)",
     "Guarda de forma permanente usuarios, cursos, publicaciones y reportes.",
     "Tamano economico (Burstable) + 32 GB", 15, 1, 1, "Base + almacen",
     "Compute ~$12 + 32 GB de espacio. Respaldo de 7 dias incluido. Sube si se agranda el servidor o el espacio."),
    ("El sitio web", "Azure Static Web Apps (Standard)",
     "La cara visible: pantallas y botones que usa la gente.",
     "Plan Standard (dominio propio + HTTPS)", 9, 1, 1, "Fijo mensual",
     "Se cobra por aplicacion. Incluye certificado de seguridad (HTTPS) sin costo extra."),
    ("La bodega de archivos", "Azure Storage (Standard LRS)",
     "Guarda fotos, videos y documentos que sube la gente.",
     "Almacenamiento estandar", 5, 1, 1, "Por uso",
     "Se paga por GB y por operaciones. Arranca barato y sube con cuanto suba la gente."),
    ("Cajas de instalacion", "Azure Container Registry (Basic)",
     "Guarda las versiones listas de cada programa para instalarlas igual siempre.",
     "Plan Basic", 5, 1, 1, "Fijo mensual",
     "Basic alcanza al inicio. Premium solo si se exige red privada del registro."),
    ("Caja fuerte de claves", "Azure Key Vault (Standard)",
     "Resguarda contrasenas y llaves; nadie las escribe en el codigo.",
     "Estandar (por operacion)", 1, 1, 1, "Por uso",
     "Costo casi nulo: se cobra por operacion (centavos)."),
    ("Camaras y monitoreo", "App Insights + Log Analytics",
     "Vigilan que todo funcione y avisan si algo falla.",
     "Por GB de registros", 5, 1, 1, "Por uso",
     "Se paga por GB de logs ingeridos. Sube con el trafico y el nivel de diagnostico."),
    ("Pasillos privados e identidades", "VNet + Managed Identity + DNS privado",
     "Comunican las areas por dentro, de forma cerrada y segura.",
     "Incluido en el diseno", 0, 1, 1, "Incluido",
     "La red y las identidades no tienen costo. El DNS privado son centavos. En modo 'max' se agregan Private Endpoints (~$7/mes c/u)."),
]

r = HEADER_ROW + 1
first_data = r
for i, (serv, prod, para, rent, precio, c1, c2, pago, cons) in enumerate(rows, start=1):
    ws.cell(row=r, column=1, value=i)
    ws.cell(row=r, column=2, value=serv)
    ws.cell(row=r, column=3, value=prod)
    ws.cell(row=r, column=4, value=para)
    ws.cell(row=r, column=5, value=rent)
    ws.cell(row=r, column=6, value=precio)
    ws.cell(row=r, column=7, value=c1)
    ws.cell(row=r, column=8, value=f"=F{r}*G{r}")
    ws.cell(row=r, column=9, value=c2)
    ws.cell(row=r, column=10, value=f"=F{r}*I{r}")
    ws.cell(row=r, column=11, value=pago)
    ws.cell(row=r, column=12, value=cons)
    for c in range(1, 13):
        cell = ws.cell(row=r, column=c)
        cell.border = border
        if c in (6, 8, 10):
            cell.font = f_bold if c in (8, 10) else f_money
            cell.number_format = MONEY
            cell.alignment = right
        elif c in (1, 7, 9):
            cell.font = f_b
            cell.alignment = center
        elif c == 2:
            cell.font = f_bold
            cell.alignment = left
        else:
            cell.font = f_b
            cell.alignment = left_top
        if i % 2 == 0:
            if cell.fill.fgColor.rgb in (None, "00000000"):
                cell.fill = fill_mist
    ws.row_dimensions[r].height = 46
    r += 1

last_data = r - 1
# Fila TOTAL
ws.cell(row=r, column=2, value="TOTAL  (USD/mes)")
ws.cell(row=r, column=8, value=f"=SUM(H{first_data}:H{last_data})")
ws.cell(row=r, column=10, value=f"=SUM(J{first_data}:J{last_data})")
for c in range(1, 13):
    cell = ws.cell(row=r, column=c)
    cell.fill = fill_total
    cell.border = border
    if c in (8, 10):
        cell.font = f_total
        cell.number_format = MONEY
        cell.alignment = right
    elif c == 2:
        cell.font = f_total
        cell.alignment = left
TOTAL_ROW = r
ws.row_dimensions[r].height = 24

# leyenda de escenarios bajo la tabla
r += 2
ws.cell(row=r, column=2, value="Version de ahorro = 1 local (App Service).  Configuracion normal = 2 locales (la que normalmente se cotiza).")
ws.cell(row=r, column=2).font = f_sub
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=12)

ws.freeze_panes = "A5"

# ============================================================
# HOJA 1: RESUMEN  (se inserta al frente)
# ============================================================
res = wb.create_sheet("Resumen", 0)
res.sheet_view.showGridLines = False
for col, w in [("A", 3), ("B", 34), ("C", 16), ("D", 34), ("E", 16)]:
    res.column_dimensions[col].width = w

res["B2"] = "HUB Digital IHP"
res["B2"].font = Font(name="Calibri", size=24, bold=True, color=SEA)
res["B3"] = "Costos de infraestructura en la nube (Azure) - resumen ejecutivo"
res["B3"].font = f_sub
res["B4"] = "Region: Mexico Central   |   Moneda: USD   |   Precios base de lista (sin IVA)"
res["B4"].font = f_b

# KPI cards (cada tarjeta ocupa 2 columnas; no se traslapan)
def kpi(lcol, rcol, row, k_text, formula, note, fill):
    res.merge_cells(f"{lcol}{row}:{rcol}{row}")
    res.merge_cells(f"{lcol}{row+1}:{rcol}{row+1}")
    res.merge_cells(f"{lcol}{row+2}:{rcol}{row+2}")
    kc = res[f"{lcol}{row}"]; kc.value = k_text; kc.font = f_kpi_k
    kc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    vc = res[f"{lcol}{row+1}"]; vc.value = formula; vc.font = f_kpi
    vc.number_format = MONEY
    vc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    nc = res[f"{lcol}{row+2}"]; nc.value = note
    nc.font = Font(name="Calibri", size=9, color=WHITE)
    nc.alignment = Alignment(horizontal="left", vertical="top", indent=1, wrap_text=True)
    for rr in (row, row+1, row+2):
        for cc in (lcol, rcol):
            res[f"{cc}{rr}"].fill = fill
    res.row_dimensions[row].height = 20
    res.row_dimensions[row+1].height = 40
    res.row_dimensions[row+2].height = 40

kpi("B", "C", 6, "VERSION DE AHORRO  (1 local)", "=Servicios!H{}".format(TOTAL_ROW),
    "Lo minimo para operar: un solo App Service compartido. Asi se arranca pagando menos.", fill_kpi_min)
kpi("D", "E", 6, "CONFIGURACION NORMAL  (2 locales)", "=Servicios!J{}".format(TOTAL_ROW),
    "La que normalmente se cotiza: recepcion en su propio local. Mas estable y segura.", fill_kpi_norm)

res["B10"] = "Diferencia = el 2o local (App Service)"
res["B10"].font = f_bold
res["C10"] = "=Servicios!J{0}-Servicios!H{0}".format(TOTAL_ROW)
res["C10"].font = f_bold
res["C10"].number_format = MONEY
res["C10"].alignment = right

# mini desglose
res["B12"] = "Desglose (configuracion normal, 2 locales)"
res["B12"].font = Font(name="Calibri", size=12, bold=True, color=SEA)
hb = res["B13"]; hb.value = "Servicio";
res["C13"] = "USD/mes"
for cellref in ("B13", "C13"):
    res[cellref].font = f_h
    res[cellref].fill = fill_head
    res[cellref].border = border
res["C13"].alignment = center
res["B13"].alignment = Alignment(horizontal="left", vertical="center")

rr = 14
for idx in range(len(rows)):
    src = first_data + idx
    res.cell(row=rr, column=2, value=f"=Servicios!B{src}")
    res.cell(row=rr, column=3, value=f"=Servicios!J{src}")
    res.cell(row=rr, column=2).font = f_b
    res.cell(row=rr, column=2).border = border
    res.cell(row=rr, column=2).alignment = Alignment(horizontal="left", vertical="center")
    res.cell(row=rr, column=3).font = f_money
    res.cell(row=rr, column=3).number_format = MONEY
    res.cell(row=rr, column=3).border = border
    res.cell(row=rr, column=3).alignment = right
    if idx % 2 == 0:
        res.cell(row=rr, column=2).fill = fill_mist
        res.cell(row=rr, column=3).fill = fill_mist
    rr += 1
res.cell(row=rr, column=2, value="TOTAL normal")
res.cell(row=rr, column=3, value=f"=Servicios!J{TOTAL_ROW}")
res.cell(row=rr, column=2).font = f_total
res.cell(row=rr, column=2).fill = fill_total
res.cell(row=rr, column=2).border = border
res.cell(row=rr, column=3).font = f_total
res.cell(row=rr, column=3).fill = fill_total
res.cell(row=rr, column=3).number_format = MONEY
res.cell(row=rr, column=3).border = border
res.cell(row=rr, column=3).alignment = right

note_r = rr + 2
res.cell(row=note_r, column=2,
         value="Nota: precios base. Con trafico real (visitas, descargas, monitoreo) la factura sube. Ver hoja 'A considerar'.")
res.cell(row=note_r, column=2).font = f_sub
res.merge_cells(start_row=note_r, start_column=2, end_row=note_r, end_column=4)

# ============================================================
# HOJA 3: A CONSIDERAR
# ============================================================
con = wb.create_sheet("A considerar")
con.sheet_view.showGridLines = False
con.column_dimensions["A"].width = 3
con.column_dimensions["B"].width = 34
con.column_dimensions["C"].width = 80

con["B2"] = "Cosas a considerar antes de contratar"
con["B2"].font = f_title
con["B3"] = "Lo que puede mover el costo hacia arriba o abajo respecto a las cifras base."
con["B3"].font = f_sub

con["B5"] = "Tema"
con["C5"] = "Que tener en cuenta"
for ref in ("B5", "C5"):
    con[ref].font = f_h
    con[ref].fill = fill_head
    con[ref].border = border
    con[ref].alignment = Alignment(horizontal="left", vertical="center")

consideraciones = [
    ("1 local vs 2 locales", "El salto mas grande de costo. 1 local (~ahorro) comparte todo; 2 locales (~normal) separa la recepcion. Es +1 App Service (~$120/mes)."),
    ("Son precios base, no la factura", "Las cifras son el piso de lista. La factura real sube con el uso: visitas, descargas, logs y transferencia de datos."),
    ("Transferencia de datos (egress)", "Sacar datos de Azure hacia internet se cobra. Con poco trafico es bajo; con video/descargas masivas crece."),
    ("Monitoreo (logs)", "App Insights y Log Analytics cobran por GB de registros. A mas trafico o mas detalle de diagnostico, mas costo."),
    ("Almacenamiento que crece", "Storage y la base de datos crecen con el uso (fotos, videos, expedientes). Se paga por lo que se ocupa."),
    ("Redis: Standard vs Basic", "Standard C1 (~$100) trae replica y garantia (SLA). Basic C1 (~$40) es mas barato pero sin respaldo: si se reinicia, se pierde la cache."),
    ("Descuentos por reserva", "Comprometer 1 o 3 anos en App Service baja el costo hasta 30%-55%. Util cuando el uso ya es estable."),
    ("Precios Dev/Test", "Entornos de prueba pueden usar tarifas Dev/Test mas baratas (no para produccion)."),
    ("Front Door: base + uso", "Standard cobra una base (~$35) mas requests y datos. Premium (con firewall WAF) sube a ~$300+/mes."),
    ("Modo de red 'max'", "Si se exige aislamiento total, se agregan Private Endpoints (~$7/mes cada uno) y mas zonas DNS. Sube el total."),
    ("Media privada (Premium)", "Servir archivos de forma privada usa Front Door Premium + storage privado: agrega varios cientos de USD/mes."),
    ("Impuestos y tipo de cambio", "Los precios son sin IVA y en USD. Para el costo final en pesos, sumar IVA y el tipo de cambio del mes."),
    ("Region Mexico Central", "Los precios pueden variar ligeramente frente a regiones de EE.UU. Confirmar en la calculadora de Azure para el numero exacto."),
    ("Lo que NO cuesta extra", "Red privada (VNet), identidades y certificados HTTPS vienen incluidos en el diseno."),
]
rr = 6
for tema, txt in consideraciones:
    con.cell(row=rr, column=2, value=tema).font = f_bold
    con.cell(row=rr, column=2).alignment = left_top
    con.cell(row=rr, column=2).border = border
    con.cell(row=rr, column=3, value=txt).font = f_b
    con.cell(row=rr, column=3).alignment = left_top
    con.cell(row=rr, column=3).border = border
    if (rr - 6) % 2 == 1:
        con.cell(row=rr, column=2).fill = fill_mist
        con.cell(row=rr, column=3).fill = fill_mist
    con.row_dimensions[rr].height = 32
    rr += 1

import os
out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Costos-HUB-Digital-IHP.xlsx")
wb.save(out)
print("OK ->", out)
